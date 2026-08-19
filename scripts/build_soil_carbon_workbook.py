# Build data/soil_carbon_calculation.xlsx -- deliverable 02 of the workshop's
# "first deliverable" list (docs/workshop.html): raw field/lab measurements ->
# carbon stock -> depth harmonisation -> the response variable step 3 models,
# with every unit conversion visible as a live formula.
#
# Source data: data/example_soil_cores.csv -- the workshop EXAMPLE dataset
# (48 slices / 16 fictional cores; see scripts/00_generate_example_data.R).
# To build a workbook from a different raw file, change SRC below.
# The instructions sheet is adapted from the WWF-Canada Blue Carbon Eelgrass
# workshop sheet that previously lived at
# "data/Carbon_DigitalData_BlankSheet - 1. Instructions.csv".
#
# Re-run:  python3 scripts/build_soil_carbon_workbook.py
#          python3 <xlsx-skill>/scripts/recalc.py data/soil_carbon_calculation.xlsx

import csv
import os

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(HERE, "data", "example_soil_cores.csv")
OUT = os.path.join(HERE, "data", "soil_carbon_calculation.xlsx")

FONT = "Arial"

# --- colour key, inherited from the source workbook -------------------------
YELLOW = PatternFill("solid", fgColor="FFF2CC")   # you type it
GREY = PatternFill("solid", fgColor="EDEDED")     # calculated -- don't type over
BAND = PatternFill("solid", fgColor="D9D9D9")     # band header strip
TITLE_FILL = PatternFill("solid", fgColor="1F3864")

BLUE_TXT = Font(name=FONT, size=10, color="1F4E79")            # field / lab entry
BLACK_TXT = Font(name=FONT, size=10, color="000000")           # calculated
GREEN_TXT = Font(name=FONT, size=10, color="006100")           # pulled from another sheet
HDR = Font(name=FONT, size=9, bold=True, color="000000")
BAND_FONT = Font(name=FONT, size=9, bold=True, color="1F3864")
TITLE_FONT = Font(name=FONT, size=12, bold=True, color="FFFFFF")
SEC_FONT = Font(name=FONT, size=10, bold=True, color="1F3864")
BODY = Font(name=FONT, size=10)
NOTE = Font(name=FONT, size=9, italic=True, color="595959")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)


def read_source():
    with open(SRC, newline="", encoding="utf-8-sig") as fh:
        rows = [r for r in csv.DictReader(fh) if r.get("Core Id")]
    out = []
    for r in rows:
        out.append(
            {
                "year": int(r["year"]),
                "core": r["Core Id"].strip(),
                "sample": r["Sample Id"].strip(),
                "lat": float(r["Latitude"]),
                "lon_supplied": float(r["Longitude"]),
                "thickness": float(r["Depth"]),
                "bd": float(r["Bulk Density"]),
                "om": float(r["OM"]),
                "soc_pct": float(r["SOC"]),
                "c_density_supplied": float(r.get("Organic Carbon Density (g/cm^3)")
                                            or r["Organic Carbon Density (g/cm^2)"]),
            }
        )
    return out


def style(cell, font, fill=None, fmt=None, align=None, border=True):
    cell.font = font
    if fill is not None:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align is not None:
        cell.alignment = align
    if border:
        cell.border = BOX


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Sheet 1 -- Instructions
# ---------------------------------------------------------------------------
def build_instructions(ws):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 42, "B": 96})

    def title(row, text):
        ws.cell(row, 1, text).font = TITLE_FONT
        for c in (1, 2):
            ws.cell(row, c).fill = TITLE_FILL
        ws.row_dimensions[row].height = 22

    def section(row, text):
        ws.cell(row, 1, text).font = SEC_FONT

    def line(row, label, text, label_font=BODY):
        a = ws.cell(row, 1, label)
        a.font = label_font
        a.alignment = WRAP_TOP
        b = ws.cell(row, 2, text)
        b.font = BODY
        b.alignment = WRAP_TOP

    r = 1
    title(r, "Workshop — Community & Regional Carbon Mapping · Ground-data sheet")
    r += 1
    line(
        r,
        "",
        "Companion to the Regional Map Updating Framework workshop (docs/workshop.html), "
        "Step 2 — Process and harmonize the ground data. Adapted from the WWF-Canada "
        "Blue Carbon Eelgrass workshop digital data sheet (Measuring Carbon in Coastal "
        "Sediments, 2026); see WHAT CHANGED below.",
    )
    r += 2

    section(r, "WHAT THIS WORKBOOK IS FOR")
    r += 1
    line(
        r,
        "",
        "Deliverable 02 of the workshop's first deliverable list: raw measurements → "
        "carbon stock → depth harmonisation → the response variable, with every unit "
        "conversion visible. Do the arithmetic here first; the R code in "
        "R/step02_harmonize_depths.R has to reproduce these numbers before it is trusted.",
    )
    r += 2

    section(r, "HOW TO USE THIS WORKBOOK")
    r += 1
    for label, text in [
        ("Sheet 2 — Core Log",
         "One row per CORE. Enter the core's position, an optional compaction factor, and "
         "landform/cover notes. Slice count, recovered length and in-situ depth calculate "
         "themselves from sheet 3."),
        ("Sheet 3 — Sample Data",
         "One row per SLICE. Enter the field columns before you leave site, then add the lab "
         "columns when results come back. Everything else calculates."),
        ("Sheet 4 — Core Summary",
         "Per-core totals and the depth-harmonised stocks. Fully automatic — nothing to "
         "type here."),
        ("Sheet 5 — R export (layers)",
         "The flat table scripts/run_02_harmonize_depths.R reads. Save as CSV to "
         "data/soil_cores_raw.csv — column names must stay exactly as they are."),
        ("Sheet 6 — R export (plots)",
         "One row per core, carrying plot_id + observed, for "
         "scripts/run_03_extract_covariates.R. Save as CSV, or build the point layer "
         "data/field_plots.gpkg from it."),
    ]:
        line(r, label, text)
        r += 1
    r += 1

    section(r, "COLOUR KEY")
    r += 1
    for label, text, fill, font in [
        ("Yellow fill / blue text", "Data you type — from the field or from the lab report.", YELLOW, BLUE_TXT),
        ("Grey fill / black text", "Calculated. Do not type over these — you will break the column.", GREY, BLACK_TXT),
        ("Green text", "Pulled from another sheet in this workbook.", None, GREEN_TXT),
    ]:
        a = ws.cell(r, 1, label)
        a.font = Font(name=FONT, size=10, bold=True, color=font.color.rgb[2:] if font.color else "000000")
        if fill:
            a.fill = fill
        ws.cell(r, 2, text).font = BODY
        ws.cell(r, 2).alignment = WRAP_TOP
        r += 1
    line(r, "", "Sheet 3 comes pre-filled with the workshop's EXAMPLE dataset -- 16 fictional "
                "cores whose numbers match the worked examples in the workshop text. Replace "
                "them with your own field data when you are ready.", NOTE)
    ws.cell(r, 2).font = NOTE
    r += 2

    section(r, "THE FIVE BANDS ON SHEET 3")
    r += 1
    for label, text in [
        ("1. Data from the field",
         "Year, core and sample id, position, measured slice thickness, coarse-fragment "
         "fraction. Copy directly from the field data sheet."),
        ("2. Calculated before the lab",
         "Slices stacked from the surface into depth-from / depth-to, then corrected to "
         "in-situ depths using the core's compaction factor."),
        ("3. Measured by the lab",
         "Dry bulk density, organic matter by loss-on-ignition, and organic carbon."),
        ("4. Calculated after the lab",
         "Carbon density and the carbon stock of each slice, in three units."),
        ("5. Harmonised to target depths",
         "Each slice's contribution to the workshop's target intervals 0–15, 15–30, "
         "30–50, 50–100 cm, split by how much of the target each slice covers — the "
         "overlap-weighting in Step 2b of the workshop."),
    ]:
        line(r, label, text)
        r += 1
    line(r, "Cross-check (cols Y–Z)",
         "The carbon density recomputed here, minus the value supplied in "
         "the raw source file. Should be 0.000000 on every row.")
    r += 2

    section(r, "THE THREE EQUATIONS")
    r += 1
    # Leading "=" would make openpyxl write these as formulas, so each reads as a
    # right-hand side with the left-hand side in the label column.
    for label, text in [
        ("Carbon density (g C/cm³)", "bulk density (g/cm³) × organic carbon (%) / 100 × (1 − coarse-fragment fraction)"),
        ("Slice stock (g C/cm²)", "carbon density (g C/cm³) × measured slice thickness (cm)"),
        ("Unit conversions", "1 g C/cm² = 10 kg C/m² = 100 Mg C/ha"),
        ("Harmonised stock", "carbon density × (overlap of the slice with the target interval, cm) × 10, summed over the core"),
    ]:
        line(r, label, text)
        r += 1
    r += 1

    section(r, "UNITS")
    r += 1
    for label, text in [
        ("Depths / thickness", "centimetres (cm)"),
        ("Dry bulk density", "g/cm³"),
        ("Organic matter, organic carbon", "percent (0–100). Some labs return a decimal fraction (0–1) — check before entering."),
        ("soc on sheet 5", "g/kg, which is simply %C × 10 — the unit soilassessment::depthharm() is given."),
        ("Carbon stock", "reported per slice as g C/cm², kg C/m² and Mg C/ha. The prior maps in "
                         "CarbonResources_Assets+Covariates are in kg C/m²; the workshop text uses Mg C/ha."),
    ]:
        line(r, label, text)
        r += 1
    r += 1

    section(r, "WHAT CHANGED FROM THE EELGRASS VERSION")
    r += 1
    for label, text in [
        ("Compaction",
         "Reduced from a computed block (corer diameter + two insertion depths) to ONE "
         "optional factor per core on sheet 2, because the community dataset arrives with "
         "bulk density already computed and no insertion depths recorded. Blank = 1.00 = no "
         "correction. The rule is unchanged: compaction corrects DEPTHS, not stocks — the "
         "dry mass in the tube already came from a taller in-situ column, so applying the "
         "factor to the stock as well double-counts it."),
        ("Corer diameter, wet/dry weights, sample volume",
         "Removed. Those columns existed to derive bulk density; here the lab reports it "
         "directly. Put them back if you start coring for bulk density yourself — and "
         "measure the corer's INTERNAL diameter with calipers, because volume sits in the "
         "denominator of bulk density and an assumed diameter biases every stock."),
        ("Depth harmonisation",
         "New band 5. The eelgrass sheet stopped at per-slice stocks; this workflow needs "
         "stocks on the workshop's fixed target intervals so cores with different slicing "
         "are comparable and can be joined to raster priors."),
        ("Units",
         "Stocks now also in kg C/m² and Mg C/ha, the units the prior maps and the "
         "workshop use."),
        ("R export sheets",
         "New sheets 5 and 6, shaped to the column names scripts/run_02 and scripts/run_03 "
         "expect."),
    ]:
        line(r, label, text)
        r += 1
    r += 1

    section(r, "THINGS THAT OFTEN CATCH PEOPLE OUT")
    r += 1
    for label, text in [
        ("Longitude signs",
         "The western hemisphere is NEGATIVE longitude. GPS units and lab sheets "
         "sometimes drop the sign; a positive longitude puts your cores on the other "
         "side of the planet, and every map step downstream fails quietly. Column D on "
         "sheet 2 is the corrected value; column E keeps the number exactly as supplied "
         "so the correction is visible."),
        ("'Depth' means slice THICKNESS here",
         "Not depth-to-bottom. Sheet 3 stacks the slices from the surface to get each "
         "slice's true depth interval (columns H-I). If your field sheet records bottom "
         "depths instead, convert before entering."),
        ("Volumetric vs areal carbon density",
         "g C per cm3 (a property of the soil) is not g C per cm2 (a stock -- density x "
         "thickness). Mixing them is a factor-of-thickness error. Sheet 3 computes the "
         "volumetric density in column Q and only turns it into a stock in column R."),
        ("Cores that stop short",
         "A core that reaches 20 cm is not a 0-30 cm observation. Do not stretch or "
         "guess -- enter it as measured. Sheet 4 flags partial coverage, and the R "
         "workflow's depth harmonization accounts for it explicitly."),
        ("%C from loss-on-ignition",
         "If your lab reports organic matter (LOI) rather than measured carbon, %C is "
         "estimated as OM x a conversion factor (cell C2 on sheet 3, default 0.58). "
         "That is an assumption, not a measurement -- record which one your numbers are."),
        ("Implausible bulk densities",
         "Dry bulk density above ~1.9 g/cm3 in soil, or above ~0.3 in peat, deserves a "
         "call to the lab before it drives an analysis: volume sits in the denominator, "
         "so one wrong volume biases every stock from that core."),
    ]:
        line(r, label, text)
        r += 1

    for row in range(1, r + 1):
        if ws.cell(row, 2).value and not ws.row_dimensions[row].height:
            ws.row_dimensions[row].height = None
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Sheet 2 -- Core Log
# ---------------------------------------------------------------------------
CORE_HEADERS = [
    ("Core Id", 12, "field"),
    ("Year", 8, "field"),
    ("Latitude (°N)", 12, "field"),
    ("Longitude (°E, corrected)", 15, "field"),
    ("Longitude as supplied", 15, "field"),
    ("Compaction factor (optional)", 14, "field"),
    ("Compaction factor used", 12, "calc"),
    ("Slices recorded", 10, "calc"),
    ("Recovered core length (cm)", 13, "calc"),
    ("In-situ core depth (cm)", 13, "calc"),
    ("Landform / cover", 22, "field"),
    ("Notes", 34, "field"),
]


def build_core_log(ws, cores, n_data):
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "Sheet 2 — Core Log  (one row per CORE)").font = Font(
        name=FONT, size=12, bold=True, color="1F3864"
    )
    ws.cell(
        2, 1,
        "Compaction factor = in-situ core length ÷ recovered length. Leave blank if you did "
        "not measure it — blank is treated as 1.00 (no correction). It corrects DEPTHS only, "
        "never stocks.",
    ).font = NOTE

    hrow = 4
    for i, (name, w, kind) in enumerate(CORE_HEADERS, start=1):
        c = ws.cell(hrow, i, name)
        style(c, HDR, BAND, align=CTR)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hrow].height = 32

    sd_core = f"'3. Sample Data'!$B$6:$B${5 + n_data}"
    sd_thick = f"'3. Sample Data'!$F$6:$F${5 + n_data}"

    for j, core in enumerate(cores):
        r = hrow + 1 + j
        style(ws.cell(r, 1, core["core"]), BLUE_TXT, YELLOW)
        style(ws.cell(r, 2, core["year"]), BLUE_TXT, YELLOW, "0")
        style(ws.cell(r, 3, core["lat"]), BLUE_TXT, YELLOW, "0.000000")
        style(ws.cell(r, 4, -abs(core["lon_supplied"])), BLUE_TXT, YELLOW, "0.000000")
        style(ws.cell(r, 5, core["lon_supplied"]), BLUE_TXT, YELLOW, "0.000000")
        style(ws.cell(r, 6, None), BLUE_TXT, YELLOW, "0.000")
        style(ws.cell(r, 7, f'=IF(F{r}="",1,F{r})'), BLACK_TXT, GREY, "0.000")
        style(ws.cell(r, 8, f"=COUNTIF({sd_core},A{r})"), GREEN_TXT, GREY, "0")
        style(ws.cell(r, 9, f"=SUMIF({sd_core},A{r},{sd_thick})"), GREEN_TXT, GREY, "0.00")
        style(ws.cell(r, 10, f"=I{r}*G{r}"), BLACK_TXT, GREY, "0.00")
        style(ws.cell(r, 11, None), BLUE_TXT, YELLOW)
        style(ws.cell(r, 12, None), BLUE_TXT, YELLOW)

    ws.cell(hrow, 4).comment = Comment(
        "Corrected longitude: western hemisphere is negative. Column E keeps the number "
        "exactly as supplied by the raw file, so any sign correction stays visible.",
        "Workshop",
    )
    ws.cell(hrow, 11).comment = Comment(
        "Fill in from field notes -- the regional model can use it as a stratifying "
        "covariate later.",
        "Workshop",
    )
    ws.freeze_panes = f"C{hrow + 1}"


# ---------------------------------------------------------------------------
# Sheet 3 -- Sample Data
# ---------------------------------------------------------------------------
BANDS = [
    ("1. Data from the field", "A", "G"),
    ("2. Calculated before the lab", "H", "L"),
    ("3. Measured by the lab", "M", "O"),
    ("4. Calculated after the lab", "P", "T"),
    ("5. Harmonised to target depths (kg C/m²)", "U", "X"),
    ("Cross-check vs. source file", "Y", "Z"),
]

SAMPLE_HEADERS = [
    ("Year", 8, "field", "0"),
    ("Core Id", 12, "field", None),
    ("Sample Id", 13, "field", None),
    ("Latitude (°N)", 11, "field", "0.000000"),
    ("Longitude (°E, corrected)", 13, "field", "0.000000"),
    ("Measured slice thickness (cm)", 12, "field", "0.000"),
    ("Coarse-fragment fraction (0–1)", 11, "field", "0.00"),
    ("Depth from, measured (cm)", 11, "calc", "0.00"),
    ("Depth to, measured (cm)", 11, "calc", "0.00"),
    ("Compaction factor", 10, "link", "0.000"),
    ("In-situ depth from (cm)", 11, "calc", "0.00"),
    ("In-situ depth to (cm)", 11, "calc", "0.00"),
    ("Dry bulk density (g/cm³)", 12, "field", "0.0000"),
    ("Organic matter, LOI (%)", 11, "field", "0.00"),
    ("Organic carbon, measured (%C)", 12, "field", "0.00"),
    ("Organic carbon used (%C)", 11, "calc", "0.00"),
    ("Carbon density (g C/cm³)", 12, "calc", "0.000000"),
    ("Slice stock (g C/cm²)", 11, "calc", "0.0000"),
    ("Slice stock (kg C/m²)", 11, "calc", "0.000"),
    ("Slice stock (Mg C/ha)", 11, "calc", "0.00"),
    ("→ 0–15 cm", 10, "calc", "0.000"),
    ("→ 15–30 cm", 10, "calc", "0.000"),
    ("→ 30–50 cm", 10, "calc", "0.000"),
    ("→ 50–100 cm", 10, "calc", "0.000"),
    ("C density as supplied (g/cm³)", 12, "field", "0.000000"),
    ("Check: this sheet − supplied", 12, "calc", "0.000000000"),
]

BAND_ROW, HDR_ROW, FIRST = 4, 5, 6
TARGETS = [(0, 15), (15, 30), (30, 50), (50, 100)]


def build_sample_data(ws, rows, n_cores):
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "Sheet 3 — Sample Data  (one row per SLICE)").font = Font(
        name=FONT, size=12, bold=True, color="1F3864"
    )

    ws.cell(2, 1, "Organic matter → organic carbon factor:").font = BODY
    f = ws.cell(2, 3, 0.58)
    style(f, BLUE_TXT, YELLOW, "0.00")
    ws.cell(2, 4,
            "van Bemmelen. Used only where 'Organic carbon, measured (%C)' (column O) is "
            "blank. In this dataset %C / OM = 0.58 on every row, so the supplied %C is "
            "itself LOI-derived, not an independent measurement.").font = NOTE
    ws.cell(3, 1,
            "Target depth intervals are the workshop's: 0–15, 15–30, 30–50, "
            "50–100 cm. Changing them means editing the four formulas in columns U–X.").font = NOTE

    for label, c0, c1 in BANDS:
        ws.merge_cells(f"{c0}{BAND_ROW}:{c1}{BAND_ROW}")
        c = ws.cell(BAND_ROW, ws[f"{c0}{BAND_ROW}"].column, label)
        style(c, BAND_FONT, BAND, align=CTR)
    for i in range(1, len(SAMPLE_HEADERS) + 1):
        ws.cell(BAND_ROW, i).border = BOX

    for i, (name, w, kind, fmt) in enumerate(SAMPLE_HEADERS, start=1):
        style(ws.cell(HDR_ROW, i, name), HDR, BAND, align=CTR)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HDR_ROW].height = 46

    cl_id = f"'2. Core Log'!$A$5:$A${4 + n_cores}"
    cl_cf = f"'2. Core Log'!$G$5:$G${4 + n_cores}"

    for j, d in enumerate(rows):
        r = FIRST + j
        first_of_core = j == 0 or rows[j - 1]["core"] != d["core"]

        style(ws.cell(r, 1, d["year"]), BLUE_TXT, YELLOW, "0")
        style(ws.cell(r, 2, d["core"]), BLUE_TXT, YELLOW)
        style(ws.cell(r, 3, d["sample"]), BLUE_TXT, YELLOW)
        style(ws.cell(r, 4, d["lat"]), BLUE_TXT, YELLOW, "0.000000")
        style(ws.cell(r, 5, -abs(d["lon_supplied"])), BLUE_TXT, YELLOW, "0.000000")
        style(ws.cell(r, 6, d["thickness"]), BLUE_TXT, YELLOW, "0.000")
        style(ws.cell(r, 7, 0), BLUE_TXT, YELLOW, "0.00")

        # band 2 -- stack slices from the surface, then correct to in-situ depth
        h = 0 if first_of_core else f"=I{r - 1}"
        style(ws.cell(r, 8, h), BLACK_TXT, GREY, "0.00")
        style(ws.cell(r, 9, f"=H{r}+F{r}"), BLACK_TXT, GREY, "0.00")
        style(ws.cell(r, 10, f"=IFERROR(INDEX({cl_cf},MATCH(B{r},{cl_id},0)),1)"), GREEN_TXT, GREY, "0.000")
        style(ws.cell(r, 11, f"=H{r}*J{r}"), BLACK_TXT, GREY, "0.00")
        style(ws.cell(r, 12, f"=I{r}*J{r}"), BLACK_TXT, GREY, "0.00")

        # band 3 -- lab
        style(ws.cell(r, 13, d["bd"]), BLUE_TXT, YELLOW, "0.0000")
        style(ws.cell(r, 14, d["om"]), BLUE_TXT, YELLOW, "0.00")
        style(ws.cell(r, 15, d["soc_pct"]), BLUE_TXT, YELLOW, "0.00")

        # band 4 -- carbon density and stock. Stock uses the MEASURED thickness:
        # the dry mass already came from a taller in-situ column.
        style(ws.cell(r, 16, f'=IF(O{r}="",N{r}*$C$2,O{r})'), BLACK_TXT, GREY, "0.00")
        style(ws.cell(r, 17, f"=M{r}*P{r}/100*(1-G{r})"), BLACK_TXT, GREY, "0.000000")
        style(ws.cell(r, 18, f"=Q{r}*F{r}"), BLACK_TXT, GREY, "0.0000")
        style(ws.cell(r, 19, f"=R{r}*10"), BLACK_TXT, GREY, "0.000")
        style(ws.cell(r, 20, f"=S{r}*10"), BLACK_TXT, GREY, "0.00")

        # band 5 -- overlap-weighted contribution to each target interval
        for k, (t0, t1) in enumerate(TARGETS):
            col = 21 + k
            style(
                ws.cell(r, col, f"=MAX(0,MIN($L{r},{t1})-MAX($K{r},{t0}))*$Q{r}*10"),
                BLACK_TXT, GREY, "0.000",
            )

        # cross-check against the carbon-density column of the raw source file
        style(ws.cell(r, 25, d["c_density_supplied"]), BLUE_TXT, YELLOW, "0.000000")
        style(ws.cell(r, 26, f"=Q{r}-Y{r}"), BLACK_TXT, GREY, "0.000000000")

    last = FIRST + len(rows) - 1
    tot = last + 1
    style(ws.cell(tot, 3, "Total / max |check|"), HDR, BAND)
    for col in (19, 20, 21, 22, 23, 24):
        fmt = "0.000" if col != 20 else "0.00"
        style(ws.cell(tot, col, f"=SUM({get_column_letter(col)}{FIRST}:{get_column_letter(col)}{last})"),
              HDR, BAND, fmt)
    # SUMPRODUCT forces array context, so this needs no Ctrl+Shift+Enter.
    style(ws.cell(tot, 26, f"=SUMPRODUCT(MAX(ABS(Z{FIRST}:Z{last})))"), HDR, BAND, "0.000000000")
    ws.cell(tot, 26).comment = Comment(
        "Largest absolute disagreement between the carbon density recomputed on this sheet "
        "and the value supplied in the raw source file, over all rows. Should be 0.",
        "Workshop",
    )
    ws.cell(tot + 1, 3,
            "Column-U–X totals are the whole-dataset stock in each target interval, not a "
            "per-core value — see sheet 4 for per-core numbers.").font = NOTE

    ws.cell(HDR_ROW, 6).comment = Comment(
        "The 'Depth' column of the raw file is the THICKNESS of each slice, not its "
        "bottom depth. Columns H-I stack the slices from the surface.",
        "Workshop",
    )
    ws.cell(HDR_ROW, 7).comment = Comment(
        "Entered as 0 (no coarse fragments), which "
        "is what the supplied carbon-density column also assumes. Change it here if the "
        "field notes say otherwise -- it feeds the stock of every row.",
        "Workshop",
    )
    ws.cell(HDR_ROW, 25).comment = Comment(
        "The volumetric carbon density supplied in the raw file: bulk density x %C / 100, "
        "in g C per cm3. It has NOT been multiplied by slice thickness -- column R is the "
        "areal stock.",
        "Workshop",
    )
    ws.freeze_panes = f"D{FIRST}"
    ws.auto_filter.ref = f"A{HDR_ROW}:Z{last}"


# ---------------------------------------------------------------------------
# Sheet 4 -- Core Summary
# ---------------------------------------------------------------------------
SUMMARY_HEADERS = [
    ("Core Id", 12), ("Year", 8), ("Latitude (°N)", 11),
    ("Longitude (°E)", 12), ("Slices", 8), ("In-situ core depth (cm)", 12),
    ("Whole-core stock (kg C/m²)", 13), ("Whole-core stock (Mg C/ha)", 13),
    ("0–15 cm (kg C/m²)", 12), ("15–30 cm (kg C/m²)", 12),
    ("30–50 cm (kg C/m²)", 12), ("50–100 cm (kg C/m²)", 12),
    ("0–30 cm (kg C/m²)", 13), ("0–30 cm (Mg C/ha)", 13),
    ("Depth coverage flag", 46),
]


def build_summary(ws, cores, n_data, n_cores):
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "Sheet 4 — Core Summary  (fully automatic — nothing to type here)").font = Font(
        name=FONT, size=12, bold=True, color="1F3864"
    )
    ws.cell(2, 1,
            "A target interval with no core material reports 'no core material', not 0. "
            "0–30 cm is the response variable step 3 extracts covariates against — use "
            "only rows flagged 'full 0–30 cm coverage'.").font = NOTE

    hrow = 4
    for i, (name, w) in enumerate(SUMMARY_HEADERS, start=1):
        style(ws.cell(hrow, i, name), HDR, BAND, align=CTR)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hrow].height = 40

    last = 5 + n_data
    sd_core = f"'3. Sample Data'!$B$6:$B${last}"
    cl_last = 4 + n_cores

    def sumif(col):
        return f"SUMIF({sd_core},$A{{r}},'3. Sample Data'!${col}$6:${col}${last})"

    for j, core in enumerate(cores):
        r = hrow + 1 + j
        cl = 5 + j  # matching row on sheet 2
        style(ws.cell(r, 1, f"='2. Core Log'!A{cl}"), GREEN_TXT, GREY)
        style(ws.cell(r, 2, f"='2. Core Log'!B{cl}"), GREEN_TXT, GREY, "0")
        style(ws.cell(r, 3, f"='2. Core Log'!C{cl}"), GREEN_TXT, GREY, "0.000000")
        style(ws.cell(r, 4, f"='2. Core Log'!D{cl}"), GREEN_TXT, GREY, "0.000000")
        style(ws.cell(r, 5, f"='2. Core Log'!H{cl}"), GREEN_TXT, GREY, "0")
        style(ws.cell(r, 6, f"='2. Core Log'!J{cl}"), GREEN_TXT, GREY, "0.00")
        style(ws.cell(r, 7, "=" + sumif("S").format(r=r)), GREEN_TXT, GREY, "0.000")
        style(ws.cell(r, 8, f"=G{r}*10"), BLACK_TXT, GREY, "0.00")

        for k, (col, (t0, _t1)) in enumerate(zip("UVWX", TARGETS)):
            c = 9 + k
            body = sumif(col).format(r=r)
            style(ws.cell(r, c, f'=IF($F{r}<={t0},"no core material",{body})'), GREEN_TXT, GREY, "0.000")

        body_015 = sumif("U").format(r=r)
        body_1530 = sumif("V").format(r=r)
        style(ws.cell(r, 13, f"={body_015}+{body_1530}"), GREEN_TXT, GREY, "0.000")
        style(ws.cell(r, 14, f"=M{r}*10"), BLACK_TXT, GREY, "0.00")
        style(
            ws.cell(r, 15,
                    f'=IF($F{r}>=30,"full 0–30 cm coverage",'
                    f'"PARTIAL — core stops at "&TEXT($F{r},"0.0")&" cm; 0–30 cm stock '
                    f'covers only "&TEXT($F{r}/30,"0%")&" of the interval")'),
            BLACK_TXT, GREY, align=WRAP_TOP,
        )

    tot = hrow + 1 + len(cores)
    style(ws.cell(tot, 1, "All cores"), HDR, BAND)
    style(ws.cell(tot, 7, f"=SUM(G{hrow+1}:G{tot-1})"), HDR, BAND, "0.000")
    style(ws.cell(tot, 13, f"=AVERAGE(M{hrow+1}:M{tot-1})"), HDR, BAND, "0.000")
    style(ws.cell(tot, 14, f"=M{tot}*10"), HDR, BAND, "0.00")
    style(ws.cell(tot, 15, "mean 0–30 cm stock over all cores, partial cores included"), NOTE, BAND)
    ws.cell(tot, 13).comment = Comment(
        "Mean across all 8 cores, including the partial-coverage ones. For the step-0 data "
        "inventory's 'ground cores' row, use the mean over full-coverage cores only.",
        "Workshop",
    )
    ws.freeze_panes = f"B{hrow + 1}"


# ---------------------------------------------------------------------------
# Sheets 5 & 6 -- R exports
# ---------------------------------------------------------------------------
def build_layer_export(ws, rows):
    headers = ["plot_id", "depth_from", "depth_to", "soc", "bulk_density", "coarse_frag"]
    for i, h in enumerate(headers, start=1):
        style(ws.cell(1, i, h), HDR, BAND, align=CTR)
        ws.column_dimensions[get_column_letter(i)].width = 14
    for j in range(len(rows)):
        src = FIRST + j
        r = 2 + j
        style(ws.cell(r, 1, f"='3. Sample Data'!B{src}"), GREEN_TXT, GREY)
        style(ws.cell(r, 2, f"='3. Sample Data'!K{src}"), GREEN_TXT, GREY, "0.00")
        style(ws.cell(r, 3, f"='3. Sample Data'!L{src}"), GREEN_TXT, GREY, "0.00")
        style(ws.cell(r, 4, f"='3. Sample Data'!P{src}*10"), GREEN_TXT, GREY, "0.0")
        style(ws.cell(r, 5, f"='3. Sample Data'!M{src}"), GREEN_TXT, GREY, "0.0000")
        style(ws.cell(r, 6, f"='3. Sample Data'!G{src}"), GREEN_TXT, GREY, "0.00")
    ws.cell(1, 1).comment = Comment(
        "Save this sheet as data/soil_cores_raw.csv -- the file "
        "scripts/run_02_harmonize_depths.R reads. Keep the header names exactly as they "
        "are: harmonize_core_depths() builds an aqp SoilProfileCollection with "
        "depths(df) <- plot_id ~ depth_from + depth_to, and depthharm() is called with "
        "var.name = 'soc'. soc is in g/kg (= %C x 10). Rows must stay sorted by plot_id "
        "with no gaps or overlaps within a core.",
        "Workshop",
    )
    ws.freeze_panes = "A2"


def build_plot_export(ws, cores):
    headers = ["plot_id", "longitude", "latitude", "observed", "observed_units",
               "depth_basis", "coverage_flag"]
    widths_ = [14, 14, 13, 12, 16, 14, 52]
    for i, (h, w) in enumerate(zip(headers, widths_), start=1):
        style(ws.cell(1, i, h), HDR, BAND, align=CTR)
        ws.column_dimensions[get_column_letter(i)].width = w
    for j in range(len(cores)):
        src = 5 + j  # sheet 4 row
        r = 2 + j
        style(ws.cell(r, 1, f"='4. Core Summary'!A{src}"), GREEN_TXT, GREY)
        style(ws.cell(r, 2, f"='4. Core Summary'!D{src}"), GREEN_TXT, GREY, "0.000000")
        style(ws.cell(r, 3, f"='4. Core Summary'!C{src}"), GREEN_TXT, GREY, "0.000000")
        style(ws.cell(r, 4, f"='4. Core Summary'!M{src}"), GREEN_TXT, GREY, "0.000")
        style(ws.cell(r, 5, "kg C/m2"), BLACK_TXT, GREY)
        style(ws.cell(r, 6, "0-30 cm"), BLACK_TXT, GREY)
        style(ws.cell(r, 7, f"='4. Core Summary'!O{src}"), GREEN_TXT, GREY, align=WRAP_TOP)
    ws.cell(1, 4).comment = Comment(
        "'observed' is the 0-30 cm stock in kg C/m2. scripts/run_03_extract_covariates.R "
        "expects plot_id + observed on a point layer (data/field_plots.gpkg). Build it from "
        "this sheet with terra::vect(df, geom = c('longitude','latitude'), crs = 'EPSG:4326') "
        "and project to a metric CRS before use. Drop or downweight rows whose coverage_flag "
        "says PARTIAL, and keep the units consistent with the prior raster you fuse against.",
        "Workshop",
    )
    ws.freeze_panes = "A2"


def main():
    rows = read_source()
    cores, seen = [], set()
    for d in rows:
        if d["core"] not in seen:
            seen.add(d["core"])
            cores.append(d)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "1. Instructions"
    build_instructions(ws1)
    build_core_log(wb.create_sheet("2. Core Log"), cores, len(rows))
    build_sample_data(wb.create_sheet("3. Sample Data"), rows, len(cores))
    build_summary(wb.create_sheet("4. Core Summary"), cores, len(rows), len(cores))
    build_layer_export(wb.create_sheet("5. R export (layers)"), rows)
    build_plot_export(wb.create_sheet("6. R export (plots)"), cores)

    wb.save(OUT)
    print(f"wrote {OUT}: {len(rows)} slices, {len(cores)} cores")


if __name__ == "__main__":
    main()
